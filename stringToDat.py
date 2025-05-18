import os
from openpyxl import load_workbook
from collections import defaultdict

WORKING_DIR = os.path.dirname(os.path.realpath(__file__))
IN_FILE = os.path.join(WORKING_DIR, "ModLanguages.xlsx")
OUT_DIR = os.path.join(WORKING_DIR, "Peasmod4", "Resources", "Languages")

def excel_to_dat(in_file):
    wb = load_workbook(in_file, read_only=True)
    language_data = defaultdict(list)

    for sheet in wb:
        sheet_name = sheet.title
        headers = [cell.value for cell in sheet[1][1:]]

        for row in sheet.iter_rows(min_row=2):
            is_empty_row = all(cell.value is None for cell in row)
            if is_empty_row:
                for lang in headers:
                    language_data[lang].append(None)
                continue

            key = row[0].value or ""
            default_value = row[1].value or "" if len(row) > 1 else ""

            for idx, cell in enumerate(row[1:]):
                if idx >= len(headers):
                    continue

                lang = headers[idx]
                cell_value = cell.value if cell.value is not None else default_value
                # 修改：不再添加工作表名称作为前缀
                full_key = key
                
                value = str(cell_value).replace('\r\n', '\n').replace('\r', '\n')
                language_data[lang].append((full_key, value))

    os.makedirs(OUT_DIR, exist_ok=True)
    for lang, entries in language_data.items():
        filename = f"{lang}.dat"
        output_path = os.path.join(OUT_DIR, filename)
        with open(output_path, 'w', encoding='utf-8') as f:
            for entry in entries:
                if entry is None:
                    f.write("\n")
                else:
                    key, value = entry
                    f.write(f'"{key}": "{value}"\n')

if __name__ == "__main__":
    if not os.path.exists(IN_FILE):
        print(f"Error: File Not Found {IN_FILE}")
        exit(1)
    
    excel_to_dat(IN_FILE)
    print(f"Done! Out: {OUT_DIR}")